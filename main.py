# -*- coding: utf-8 -*-

import time
import openpyxl  # 엑셀파일을 다루는 모듈
import discord
import datetime
import tz
import config
import asyncio
import re
import os
import random
import pytz
import os

from PIL import ImageGrab
import pyautogui as pag

intents = discord.Intents.all()
client = discord.Client(intents=intents)

dev = '충북대 코딩냥이가 전한다냥'

auto_self_diagnosis_list = []
self_diagnosis_list = []
notify_at_8am = []

done = False


@client.event
async def on_guild_join(guild):
    def check(event):
        return event.target.id == client.user.id

    bot_entry = await guild.audit_logs(action=discord.AuditLogAction.bot_add).find(check)
    embed = discord.Embed(title='코냥이', timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x754D34)
    embed.add_field(name="안녕하세요? 저는 **코냥이** 입니다.",
                    value='\n명령어를 보고 싶으시다면 ``코냥아 도움말`` 을 입력하세요.',
                    inline=False)
    embed.set_footer(text=f'{guild.name} 서버에 초대해주셔서 정말로 감사합니다!')

    try:
        await bot_entry.user.send(embed=embed)
    except:
        pass


@client.event
async def on_error(event, *args, **kwargs):
    pass


@client.event
async def on_member_join(member):
    embed = discord.Embed(title='코냥이', timestamp=datetime.datetime.now(pytz.timezone('UTC')), color=0x754D34)
    embed.add_field(name="안녕하세요? 저는 **코냥이** 입니다.",
                    value='\n명령어를 보고 싶으시다면 ``코냥아 도움말`` 을 입력하세요.',
                    inline=False)
    embed.set_footer(text=dev)

    await member.send(embed=embed)


@client.event
async def on_error(event, *args, **kwargs):
    pass


@client.event
async def on_connect():
    await client.change_presence(activity=discord.Activity(type=discord.ActivityType.listening, name=f'코냥아 도움말'))
    print(f'{client.user.name} (으)로 접속됨.')
    print('─────────────────────────────────────────────────────────────────────────────────')


@client.event
async def my_background_task():
    await client.wait_until_ready()
    while True:
        try:
            await client.change_presence(
                activity=discord.Activity(type=discord.ActivityType.listening, name=f'코냥아 도움말'))
        except:
            pass
        await asyncio.sleep(5)
        try:
            await client.change_presence(
                activity=discord.Activity(type=discord.ActivityType.listening, name=f'{len(client.guilds)}개의 서버에서'))
        except:
            pass
        await asyncio.sleep(5)
        try:
            await client.change_presence(
                activity=discord.Activity(type=discord.ActivityType.listening,
                                          name=f'{len(list(filter(lambda x: not (x.bot), client.users)))}명의 명령어를'))
        except:
            pass
        await asyncio.sleep(5)


@client.event
async def self_diagnosis_task():
    await client.wait_until_ready()

    while True:
        if self_diagnosis_list:
            d = datetime.datetime.now()

            if not os.path.exists(f"./2022/{d.month}"):
                os.makedirs(f"./2022/{d.month}")

            if not os.path.exists(f"./2022/{d.month}/{d.day}"):
                os.makedirs(f"./2022/{d.month}/{d.day}")

            if not os.path.exists(f"./2022/{d.month}/{d.day}/self_diagnosis"):
                os.makedirs(f"./2022/{d.month}/{d.day}/self_diagnosis")

            start = time.time()
            data = self_diagnosis_list.pop(0)
            path = f"./2022/{d.month}/{d.day}/self_diagnosis/{data[2].name}.png"

            pag.click((460, 1050))  # 크롬 누르기
            pag.click((957, 394))  # 개신누리 아이디 누르기
            pag.typewrite(data[0])  # 개신누리 아이디 타이핑
            pag.press('\t')  # tap키 타이핑
            pag.typewrite(data[1])  # 개신누리 비밀번호 타이핑
            pag.press('enter')  # enter키 타이핑
            await asyncio.sleep(5)

            if pag.locateOnScreen('로그인 오류.PNG'):

                embed = discord.Embed(title="아이디 혹은 비밀번호가 일치하지 않습니다.",
                                      description=f"**코냥아 설정** 을 통하여 아이디와 비밀번호를 다시 설정해주세요",
                                      timestamp=datetime.datetime.now(pytz.timezone('UTC')),
                                      colour=0x754D34)
                embed.set_author(name=f"{data[2].name} 님의 로그인 오류", icon_url=data[2].avatar_url)
                embed.set_footer(text=dev)
                await data[2].send(embed=embed)

                file = openpyxl.load_workbook('자가진단.xlsx')
                sheet = file.active
                work = f'{data[2].id}'
                for i in range(1, 1000):
                    if str(sheet['A' + str(i)].value) == str(work):
                        id = str(sheet['B' + str(i)].value)
                        psw = str(sheet['C' + str(i)].value)
                        sheet['A' + str(i)].value = '-'
                        sheet['B' + str(i)].value = ''
                        sheet['C' + str(i)].value = ''
                        break
                try:
                    print(f'-ㅣ코냥아 데이터 삭제ㅣ{data[2].name} ( {id} {psw} )\n')
                    file.save('자가진단.xlsx')
                    embed = discord.Embed(description=f"로그인 오류로 인해, 자가진단 데이터에 저장된 회원정보를 제거합니다.",
                                          timestamp=datetime.datetime.now(pytz.timezone('UTC')), colour=0x754D34)
                    embed.set_author(name=f"{data[2].name} 님의 회원정보 삭제", icon_url=data[2].avatar_url)
                    embed.set_footer(text=dev)
                    await data[2].send(embed=embed)
                except:
                    pass
                try:
                    pag.click(pag.center(pag.locateOnScreen('okay.PNG')))
                except:
                    pass
                pag.hotkey('F5')
                await asyncio.sleep(1)
                pag.click((1136, 225))  # 팝업차단 창 누르기
                pag.click((815, 1050))  # 파이참 누르기
                continue

            screen = ImageGrab.grab()  # 스크린샷 찍기
            rgb = screen.getpixel((1050, 510))  # 색 추출하여 중복로그인상태인지 확인하기

            if rgb == (231, 231, 231):
                pag.click((926, 620))  # 여기서 로그인하기 클릭하기
                await asyncio.sleep(10)

            elif pag.locateOnScreen('교직원 조회.PNG') is None:
                embed = discord.Embed(title="자가진단과정에서 오류가 발생했습니다",
                                      description=f"잠시후에 자가진단을 다시 실시해주세요",
                                      timestamp=datetime.datetime.now(pytz.timezone('UTC')),
                                      colour=0x754D34)
                embed.set_author(name=f"{data[2].name} 님의 자가진단 오류", icon_url=data[2].avatar_url)
                embed.set_footer(text=dev)
                await data[2].send(embed=embed)

                pag.hotkey('ctrl', 't')
                await asyncio.sleep(1)
                pag.typewrite('https://eis.cbnu.ac.kr/')
                pag.press('enter')
                await asyncio.sleep(1)
                pag.click((115, 16))
                pag.hotkey('ctrl', 'w')
                await asyncio.sleep(1)
                pag.click((1136, 225))  # 팝업차단 창 누르기
                pag.click((815, 1050))  # 파이참 누르기
                continue

            pag.click((92, 370))  # 자가진단 메뉴 누르기
            await asyncio.sleep(1)
            screen = ImageGrab.grab()  # 스크린샷 찍기
            rgb = screen.getpixel((876, 552))  # 색 추출하여 중복로그인상태인지 확인하기

            if rgb == (255, 129, 28):
                pag.click((960, 620))  # 여기서 로그인하기 클릭하기
                await asyncio.sleep(1)
                pag.click((1136, 225))  # 팝업차단 창 누르기
                pag.click((815, 1050))  # 파이참 누르기
                embed = discord.Embed(title="자가진단 실행중에 사용자가 로그하였습니다",
                                      description=f"잠시후에 자가진단을 다시 실시해주세요",
                                      timestamp=datetime.datetime.now(pytz.timezone('UTC')),
                                      colour=0x754D34)
                embed.set_author(name=f"{data[2].name} 님의 로그인 오류", icon_url=data[2].avatar_url)
                embed.set_footer(text=dev)
                await data[2].send(embed=embed)
                continue

            await asyncio.sleep(5)  # 로딩 기다리기

            pag.click((310, 453))  # 진단하기 버튼 누르기
            await asyncio.sleep(1)
            screen = ImageGrab.grab()  # 스크린샷 찍기
            rgb = screen.getpixel((876, 552))  # 색 추출하여 중복로그인상태인지 확인하기

            if rgb == (255, 129, 28):
                pag.click((960, 620))  # 여기서 로그인하기 클릭하기
                await asyncio.sleep(1)
                pag.click((1136, 225))  # 팝업차단 창 누르기
                pag.click((815, 1050))  # 파이참 누르기
                embed = discord.Embed(title="자가진단 실행중에 사용자가 로그하였습니다",
                                      description=f"잠시후에 자가진단을 다시 실시해주세요",
                                      timestamp=datetime.datetime.now(pytz.timezone('UTC')),
                                      colour=0x754D34)
                embed.set_author(name=f"{data[2].name} 님의 로그인 오류", icon_url=data[2].avatar_url)
                embed.set_footer(text=dev)
                await data[2].send(embed=embed)
                continue

            await asyncio.sleep(5)  # 로딩 기다리기

            pag.click((937, 495))  # 1번 문항 아니오
            pag.click((937, 575))  # 2번 문항 아니오
            # pag.moveTo((1307, 452))  # 스크롤 위에 마우스커서 이동
            # pag.dragTo(x=1271, y=765, duration=0.5)  # 스크롤
            pag.click((1271, 765))  # 스크롤
            pag.click((937, 486))  # 3번 문항 아니오
            pag.click((937, 564))  # 4번 문항 아니오
            pag.click((937, 644))  # 5번 문항 아니오
            pag.click((937, 723))  # 6번 문항 아니오
            pag.click((957, 779))  # 제출 버튼 누르기
            await asyncio.sleep(3)  # 로딩 기다리기

            pag.click((933, 618))  # 예 버튼 누르기
            await asyncio.sleep(1)
            screen = ImageGrab.grab()  # 스크린샷 찍기
            rgb = screen.getpixel((876, 552))  # 색 추출하여 중복로그인상태인지 확인하기

            if rgb == (255, 129, 28):
                pag.click((960, 620))  # 여기서 로그인하기 클릭하기
                await asyncio.sleep(1)
                pag.click((1136, 225))  # 팝업차단 창 누르기
                pag.click((815, 1050))  # 파이참 누르기
                embed = discord.Embed(title="자가진단 실행중에 사용자가 로그하였습니다",
                                      description=f"잠시후에 자가진단을 다시 실시해주세요",
                                      timestamp=datetime.datetime.now(pytz.timezone('UTC')),
                                      colour=0x754D34)
                embed.set_author(name=f"{data[2].name} 님의 로그인 오류", icon_url=data[2].avatar_url)
                embed.set_footer(text=dev)
                await data[2].send(embed=embed)
                continue

            await asyncio.sleep(5)  # 로딩 기다리기

            img = ImageGrab.grab([768, 343, 1160, 609])
            img.save(path)

            pag.click((1168, 313))  # X버튼 누르기
            await asyncio.sleep(1)
            screen = ImageGrab.grab()  # 스크린샷 찍기
            rgb = screen.getpixel((876, 552))  # 색 추출하여 중복로그인상태인지 확인하기

            if rgb == (255, 129, 28):
                pag.click((960, 620))  # 여기서 로그인하기 클릭하기
                await asyncio.sleep(1)
                pag.click((1136, 225))  # 팝업차단 창 누르기
                pag.click((815, 1050))  # 파이참 누르기
                embed = discord.Embed(title="자가진단 실행중에 사용자가 로그인하였습니다",
                                      description=f"잠시후에 자가진단을 다시 실시해주세요",
                                      timestamp=datetime.datetime.now(pytz.timezone('UTC')),
                                      colour=0x754D34)
                embed.set_author(name=f"{data[2].name} 님의 로그인 오류", icon_url=data[2].avatar_url)
                embed.set_footer(text=dev)
                await data[2].send(embed=embed)
                continue
            pag.click((1868, 125))  # 로그아웃 누르기
            await asyncio.sleep(1)

            screen = ImageGrab.grab()  # 스크린샷 찍기
            rgb = screen.getpixel((876, 552))  # 색 추출하여 중복로그인상태인지 확인하기

            if rgb == (255, 129, 28):
                pag.click((960, 620))  # 여기서 로그인하기 클릭하기
                await asyncio.sleep(1)
                pag.click((1136, 225))  # 팝업차단 창 누르기
                pag.click((815, 1050))  # 파이참 누르기
                embed = discord.Embed(title="자가진단 실행중에 사용자가 로그하였습니다",
                                      description=f"잠시후에 자가진단을 다시 실시해주세요",
                                      timestamp=datetime.datetime.now(pytz.timezone('UTC')),
                                      colour=0x754D34)
                embed.set_author(name=f"{data[2].name} 님의 로그인 오류", icon_url=data[2].avatar_url)
                embed.set_footer(text=dev)
                await data[2].send(embed=embed)
                continue
            await asyncio.sleep(1)
            pag.click((1136, 225))  # 팝업차단 창 누르기
            pag.click((815, 1050))  # 파이참 누르기

            image = discord.File(f"{path}", filename="image.png")

            total = (time.time() - start)
            embed = discord.Embed(title="자가진단을 완료하였습니다",
                                  description=f"[ 걸린 시간 : {round(total)}초 ]\n\n**- 공지\n이 스크립트는 건강상의 문제가 없을 경우, 브라우저를 "
                                              f"열고 복잡한 인증 절차를 거칠 필요 없이 하나의 명령어로 빠르게 자가진단을 마치기 위해서 개발되었습니다. 실행 전 반드시 "
                                              f"개인 건강상태를 확인해주시길 바랍니다.**\n\n- 혹여나 유증상인데 이미 앱에서 무증상으로 제출했다면 자가진단 홈페이지에 "
                                              f"직접 접속해서 다시 제출하시길 바랍니다.",
                                  timestamp=datetime.datetime.now(pytz.timezone('UTC')),
                                  colour=0x754D34)
            embed.set_image(url="attachment://image.png")
            embed.set_author(name=f"{data[2].name} 님의 자가진단", icon_url=data[2].avatar_url)
            embed.set_footer(text=dev)
            await data[2].send("자가진단을 완료하였습니다", embed=embed, file=image)

            print('#ㅣ코냥아 자가진단ㅣ' + f'{data[2].name}\n')

            await asyncio.sleep(1)

        elif auto_self_diagnosis_list:
            start = time.time()

            d = datetime.datetime.now()

            if not os.path.exists(f"./2022/{d.month}"):
                os.makedirs(f"./2022/{d.month}")

            if not os.path.exists(f"./2022/{d.month}/{d.day}"):
                os.makedirs(f"./2022/{d.month}/{d.day}")

            if not os.path.exists(f"./2022/{d.month}/{d.day}/auto_self_diagnosis"):
                os.makedirs(f"./2022/{d.month}/{d.day}/auto_self_diagnosis")

            data = auto_self_diagnosis_list.pop(0)
            path = f"./2022/{d.month}/{d.day}/auto_self_diagnosis/{data[2].name}.png"

            pag.click((460, 1050))  # 크롬 누르기
            pag.click((957, 394))  # 개신누리 아이디 누르기
            pag.typewrite(data[0])  # 개신누리 아이디 타이핑
            pag.press('\t')  # tap키 타이핑
            pag.typewrite(data[1])  # 개신누리 비밀번호 타이핑
            pag.press('enter')  # enter키 타이핑
            await asyncio.sleep(5)

            if pag.locateOnScreen('로그인 오류.PNG'):
                type = "login_error"
                notify_at_8am.append((type, data[2]))

                file = openpyxl.load_workbook('자가진단.xlsx')
                sheet = file.active
                work = f'{data[2].id}'
                for i in range(1, 1000):
                    if str(sheet['A' + str(i)].value) == str(work):
                        id = str(sheet['B' + str(i)].value)
                        psw = str(sheet['C' + str(i)].value)
                        sheet['A' + str(i)].value = '-'
                        sheet['B' + str(i)].value = ''
                        sheet['C' + str(i)].value = ''
                        break
                try:
                    print(f'-ㅣ코냥아 데이터 삭제ㅣ{data[2].name} ( {id} {psw} )\n')
                    file.save('자가진단.xlsx')
                    type = "remove_data"
                    notify_at_8am.append((type, data[2]))
                except:
                    pass
                pag.click(pag.center(pag.locateOnScreen('okay.PNG')))
                pag.hotkey('F5')
                await asyncio.sleep(1)
                pag.click((1136, 225))  # 팝업차단 창 누르기
                pag.click((815, 1050))  # 파이참 누르기
                continue

            screen = ImageGrab.grab()  # 스크린샷 찍기
            rgb = screen.getpixel((1050, 510))  # 색 추출하여 중복로그인상태인지 확인하기

            if rgb == (231, 231, 231):
                pag.click((926, 620))  # 여기서 로그인하기 클릭하기
                await asyncio.sleep(10)

            elif pag.locateOnScreen('교직원 조회.PNG') is None:
                type = "error"
                notify_at_8am.append((type, data[2]))

                pag.hotkey('ctrl', 't')
                await asyncio.sleep(1)
                pag.typewrite('https://eis.cbnu.ac.kr/')
                pag.press('enter')
                await asyncio.sleep(1)
                pag.click((115, 16))
                pag.hotkey('ctrl', 'w')
                await asyncio.sleep(1)
                pag.click((1136, 225))  # 팝업차단 창 누르기
                pag.click((815, 1050))  # 파이참 누르기
                continue

            pag.click((92, 370))  # 자가진단 메뉴 누르기
            await asyncio.sleep(1)
            screen = ImageGrab.grab()  # 스크린샷 찍기
            rgb = screen.getpixel((876, 552))  # 색 추출하여 중복로그인상태인지 확인하기

            if rgb == (255, 129, 28):
                pag.click((960, 620))  # 여기서 로그인하기 클릭하기
                await asyncio.sleep(1)
                pag.click((1136, 225))  # 팝업차단 창 누르기
                pag.click((815, 1050))  # 파이참 누르기
                type = "session_expired"
                notify_at_8am.append((type, data[2]))
                continue

            await asyncio.sleep(5)  # 로딩 기다리기

            pag.click((310, 453))  # 진단하기 버튼 누르기
            await asyncio.sleep(1)
            screen = ImageGrab.grab()  # 스크린샷 찍기
            rgb = screen.getpixel((876, 552))  # 색 추출하여 중복로그인상태인지 확인하기

            if rgb == (255, 129, 28):
                pag.click((960, 620))  # 여기서 로그인하기 클릭하기
                await asyncio.sleep(1)
                pag.click((1136, 225))  # 팝업차단 창 누르기
                pag.click((815, 1050))  # 파이참 누르기
                type = "session_expired"
                notify_at_8am.append((type, data[2]))
                continue

            await asyncio.sleep(5)  # 로딩 기다리기

            pag.click((937, 495))  # 1번 문항 아니오
            pag.click((937, 575))  # 2번 문항 아니오
            # pag.moveTo((1307, 452))  # 스크롤 위에 마우스커서 이동
            # pag.dragTo(x=1271, y=765, duration=0.5)  # 스크롤
            pag.click((1271, 765))  # 스크롤
            pag.click((937, 486))  # 3번 문항 아니오
            pag.click((937, 564))  # 4번 문항 아니오
            pag.click((937, 644))  # 5번 문항 아니오
            pag.click((937, 723))  # 6번 문항 아니오
            pag.click((957, 779))  # 제출 버튼 누르기
            await asyncio.sleep(3)  # 로딩 기다리기

            pag.click((933, 618))  # 예 버튼 누르기
            await asyncio.sleep(1)
            screen = ImageGrab.grab()  # 스크린샷 찍기
            rgb = screen.getpixel((876, 552))  # 색 추출하여 중복로그인상태인지 확인하기

            if rgb == (255, 129, 28):
                pag.click((960, 620))  # 여기서 로그인하기 클릭하기
                await asyncio.sleep(1)
                pag.click((1136, 225))  # 팝업차단 창 누르기
                pag.click((815, 1050))  # 파이참 누르기
                type = "session_expired"
                notify_at_8am.append((type, data[2]))
                continue

            await asyncio.sleep(5)  # 로딩 기다리기

            img = ImageGrab.grab([768, 343, 1160, 609])
            img.save(path)

            pag.click((1168, 313))  # X버튼 누르기
            await asyncio.sleep(1)
            screen = ImageGrab.grab()  # 스크린샷 찍기
            rgb = screen.getpixel((876, 552))  # 색 추출하여 중복로그인상태인지 확인하기

            if rgb == (255, 129, 28):
                pag.click((960, 620))  # 여기서 로그인하기 클릭하기
                await asyncio.sleep(1)
                pag.click((1136, 225))  # 팝업차단 창 누르기
                pag.click((815, 1050))  # 파이참 누르기
                type = "session_expired"
                notify_at_8am.append((type, data[2]))
                continue
            pag.click((1868, 125))  # 로그아웃 누르기
            await asyncio.sleep(1)

            screen = ImageGrab.grab()  # 스크린샷 찍기
            rgb = screen.getpixel((876, 552))  # 색 추출하여 중복로그인상태인지 확인하기

            if rgb == (255, 129, 28):
                pag.click((960, 620))  # 여기서 로그인하기 클릭하기
                await asyncio.sleep(1)
                pag.click((1136, 225))  # 팝업차단 창 누르기
                pag.click((815, 1050))  # 파이참 누르기
                type = "session_expired"
                notify_at_8am.append((type, data[2]))
                continue
            await asyncio.sleep(1)
            pag.click((1136, 225))  # 팝업차단 창 누르기
            pag.click((815, 1050))  # 파이참 누르기

            image = discord.File(f"{path}", filename="image.png")

            total = (time.time() - start)
            type = "success"
            notify_at_8am.append((type, total, data[2], image))

            print('#ㅣ코냥아 자동자가진단ㅣ' + f'{data[2].name}\n')

            await asyncio.sleep(1)

        else:
            await asyncio.sleep(10)


@client.event
async def auto_self_diagnosis():
    await client.wait_until_ready()
    global done

    while True:
        if done and datetime.datetime.now().hour == 23:
            done = False

        if (not done) and (not auto_self_diagnosis_list) and datetime.datetime.now().hour == 6:
            f = open("자동자가진단.txt", 'r')
            lines = f.read().splitlines()

            for line in lines:
                file = openpyxl.load_workbook('자가진단.xlsx')
                sheet = file.active
                work = f'{line}'
                state = False

                for i in range(1, 1000):
                    if str(sheet['A' + str(i)].value) == str(work):
                        id = str(sheet['B' + str(i)].value)
                        psw = str(sheet['C' + str(i)].value)
                        state = True
                        break

                try:
                    if state and client.get_user(int(line)):
                        auto_self_diagnosis_list.append((id, psw, client.get_user(int(line))))
                    else:
                        pass

                except UnboundLocalError:
                    type = "cantfound"
                    notify_at_8am.append((type, client.get_user(int(line))))

            done = True

            await asyncio.sleep(1)

        else:
            await asyncio.sleep(10)


@client.event
async def notify():
    await client.wait_until_ready()

    while True:
        if notify_at_8am and (datetime.datetime.now().hour >= 8):
            data = notify_at_8am.pop(0)

            if data[0] == "success":
                embed = discord.Embed(title="자동자가진단이 완료되었습니다.",
                                      description=f"[ 걸린 시간 : {round(data[1])}초 ]\n\n**- 공지\n이 스크립트는 건강상의 문제가 없을 경우, 브라우저를 "
                                                  f"열고 복잡한 인증 절차를 거칠 필요 없이 하나의 명령어로 빠르게 자가진단을 마치기 위해서 개발되었습니다. 실행 전 반드시 "
                                                  f"개인 건강상태를 확인해주시길 바랍니다.**\n\n- 혹여나 유증상인데 이미 앱에서 무증상으로 제출했다면 자가진단 홈페이지에 "
                                                  f"직접 접속해서 다시 제출하시길 바랍니다.",
                                      timestamp=datetime.datetime.now(pytz.timezone('UTC')),
                                      colour=0x754D34)
                embed.set_image(url="attachment://image.png")
                embed.set_author(name=f"{data[2].name} 님의 자동자가진단", icon_url=data[2].avatar_url)
                embed.set_footer(text=dev)
                await data[2].send("자가진단을 완료하였습니다", embed=embed, file=data[3])

            elif data[0] == "cantfound":
                embed = discord.Embed(description=f"회원의 정보를 찾을 수가 없습니다, ``코냥아 설정`` 으로 설정해 주세요.",
                                      timestamp=datetime.datetime.now(pytz.timezone('UTC')), colour=0x754D34)
                embed.set_author(name=f"{data[1]} 님의 회원탈퇴", icon_url=data[1].avatar_url)
                embed.set_footer(text=dev)
                await data[1].send(embed=embed)

            elif data[0] == "login_error":
                embed = discord.Embed(title="아이디 혹은 비밀번호가 일치하지 않습니다.",
                                      description=f"**코냥아 설정** 을 통하여 아이디와 비밀번호를 다시 설정해주세요",
                                      timestamp=datetime.datetime.now(pytz.timezone('UTC')),
                                      colour=0x754D34)
                embed.set_author(name=f"{data[1].name} 님의 로그인 오류", icon_url=data[1].avatar_url)
                embed.set_footer(text=dev)
                await data[1].send(embed=embed)

            elif data[0] == "error":
                embed = discord.Embed(title="자가진단과정에서 오류가 발생했습니다",
                                      description=f"잠시후에 자가진단을 다시 실시해주세요",
                                      timestamp=datetime.datetime.now(pytz.timezone('UTC')),
                                      colour=0x754D34)
                embed.set_author(name=f"{data[1].name} 님의 자가진단 오류", icon_url=data[1].avatar_url)
                embed.set_footer(text=dev)
                await data[1].send(embed=embed)

            elif data[0] == "session_expired":
                embed = discord.Embed(title="자가진단 실행중에 사용자가 로그하였습니다",
                                      description=f"잠시후에 자가진단을 다시 실시해주세요",
                                      timestamp=datetime.datetime.now(pytz.timezone('UTC')),
                                      colour=0x754D34)
                embed.set_author(name=f"{data[1].name} 님의 로그인 오류", icon_url=data[1].avatar_url)
                embed.set_footer(text=dev)
                await data[1].send(embed=embed)

            elif data[0] == "remove_data":
                embed = discord.Embed(description=f"로그인 오류로 인해, 자가진단 데이터에 저장된 회원정보를 제거합니다.",
                                      timestamp=datetime.datetime.now(pytz.timezone('UTC')), colour=0x754D34)
                embed.set_author(name=f"{data[1].name} 님의 회원정보 삭제", icon_url=data[1].avatar_url)
                embed.set_footer(text=dev)
                await data[1].send(embed=embed)

            await asyncio.sleep(1)

        else:
            await asyncio.sleep(10)


@client.event
async def on_message(message):
    if message.content == '코냥아 탈퇴':
        a = "코냥아 승인"
        timeout = 15
        embed = discord.Embed(description=f"회원 탈퇴를 원하시면 15초 이내에 코냥아 승인을 입력해 주세요.",
                              timestamp=datetime.datetime.now(pytz.timezone('UTC')), colour=0x754D34)
        embed.set_author(name=f"{message.author.name} 님의 회원탈퇴", icon_url=message.author.avatar_url)
        embed.set_footer(text=dev)
        await message.reply(embed=embed, mention_author=True)

        def check(ylt):
            return ylt.author == message.author and ylt.channel == message.channel

        try:
            ylt = await client.wait_for("message", timeout=timeout, check=check)
        except:
            embed = discord.Embed(description=f"시간을 초과하였습니다, 회원 탈퇴가 거절됩니다.",
                                  timestamp=datetime.datetime.now(pytz.timezone('UTC')), colour=0x754D34)
            embed.set_author(name=f"{message.author.name} 님의 회원탈퇴", icon_url=message.author.avatar_url)
            embed.set_footer(text=dev)
            await message.reply(embed=embed, mention_author=True)

        if ylt.content == a:
            file = openpyxl.load_workbook('자가진단.xlsx')
            sheet = file.active
            work = f'{message.author.id}'
            for i in range(1, 1000):
                if str(sheet['A' + str(i)].value) == str(work):
                    id = str(sheet['B' + str(i)].value)
                    psw = str(sheet['C' + str(i)].value)
                    sheet['A' + str(i)].value = '-'
                    sheet['B' + str(i)].value = ''
                    sheet['C' + str(i)].value = ''
                    break
            try:
                print(f'-ㅣ코냥아 회원탈퇴ㅣ{message.author.name} ( {id} {psw} )\n')
            except UnboundLocalError:
                embed = discord.Embed(description=f"회원의 정보를 찾을 수가 없습니다, ``코냥아 설정`` 으로 설정해 주세요.",
                                      timestamp=datetime.datetime.now(pytz.timezone('UTC')), colour=0x754D34)
                embed.set_author(name=f"{message.author.name} 님의 회원탈퇴", icon_url=message.author.avatar_url)
                embed.set_footer(text=dev)
                await message.reply(embed=embed, mention_author=True)
                await message.author.send(embed=embed)
            else:
                file.save('자가진단.xlsx')
                embed = discord.Embed(description=f"회원 탈퇴로 인해, 자가진단 데이터에 저장된 회원정보를 제거합니다.",
                                      timestamp=datetime.datetime.now(pytz.timezone('UTC')), colour=0x754D34)
                embed.set_author(name=f"{message.author.name} 님의 회원탈퇴", icon_url=message.author.avatar_url)
                embed.set_footer(text=dev)
                await message.reply(embed=embed, mention_author=True)

            text = open('자동자가진단.txt', 'r')
            lines = text.read().splitlines()
            text.close()

            text = open('자동자가진단.txt', 'w')
            for line in lines:
                if str(message.author.id) == str(line):
                    continue
                text.write(line + '\n')
            text.close()

            text = open('이용약관.txt', 'r')
            lines = text.read().splitlines()
            text.close()

            text = open('이용약관.txt', 'w')
            for line in lines:
                if str(message.author.id) == str(line):
                    continue
                text.write(line + '\n')
            text.close()

        else:
            embed = discord.Embed(description=f"승인을 받지 못하였습니다, 회원 탈퇴가 거절됩니다.",
                                  timestamp=datetime.datetime.now(pytz.timezone('UTC')), colour=0x754D34)
            embed.set_author(name=f"{message.author.name} 님의 회원탈퇴", icon_url=message.author.avatar_url)
            embed.set_footer(text=dev)
            await message.reply(embed=embed, mention_author=True)

    if message.content == '코냥아 자가진단':

        person = message.author

        text1 = open('이용약관.txt', 'r')
        text1 = text1.read()

        if str(message.author.id) in str(text1):
            file = openpyxl.load_workbook('자가진단.xlsx')
            sheet = file.active
            work = f'{message.author.id}'
            for i in range(1, 1000):
                if str(sheet['A' + str(i)].value) == str(work):
                    id = str(sheet['B' + str(i)].value)
                    psw = str(sheet['C' + str(i)].value)
                    break

            else:
                embed = discord.Embed(description=f"회원의 정보를 찾을 수가 없습니다, ``코냥아 설정`` 으로 설정해 주세요.",
                                      timestamp=datetime.datetime.now(pytz.timezone('UTC')), colour=0x754D34)
                embed.set_author(name=f"{message.author.name} 님의 자가진단", icon_url=message.author.avatar_url)
                embed.set_footer(text=dev)
                await message.reply(embed=embed, mention_author=True)
                return

            self_diagnosis_list.append((id, psw, person))

        else:
            print('#ㅣ코냥아 자가진단ㅣ' + f'{message.author.name}, 이용약관에 동의하셔야 합니다.\n')
            embed = discord.Embed(description=f"자가진단 서비스를 이용 하실려면 이용약관에 동의하셔야 합니다.",
                                  timestamp=datetime.datetime.now(pytz.timezone('UTC')), colour=0x754D34)
            embed.set_author(name=f"{message.author.name} 님의 자가진단", icon_url=message.author.avatar_url)
            embed.set_footer(text=dev)
            await message.reply(embed=embed, mention_author=True)

    if message.content.startswith('코냥아 설정'):
        text1 = open('이용약관.txt', 'r')
        text1 = text1.read()
        if str(message.author.id) in str(text1):
            if message.channel.type is discord.ChannelType.private:
                try:
                    u = message.content[6:].split(' ')
                    file = openpyxl.load_workbook('자가진단.xlsx')
                    sheet = file.active
                    for i in range(1, 1000):
                        if sheet['A' + str(i)].value == '-' or sheet['A' + str(i)].value == str(message.author.id):
                            sheet['A' + str(i)].value = str(message.author.id)
                            sheet['B' + str(i)].value = str(u[1])
                            sheet['C' + str(i)].value = str(u[2])
                            break

                    file.save('자가진단.xlsx')
                    embed = discord.Embed(
                        description=f"**- 공지\n이 스크립트는 건강상의 문제가 없을 경우, 브라우저를 열고 복잡한 인증 절차를 거칠 필요 없이 하나의 명령어로 빠르게 자가진단을 "
                                    f"마치기 위해서 개발되었습니다. 실행 전 반드시 개인 건강상태를 확인해주시길 바랍니다.**\n\n- 혹여나 유증상인데 이미 앱에서 무증상으로 "
                                    f"제출했다면 자가진단 홈페이지에 직접 접속해서 다시 제출하시길 바랍니다. ",
                        timestamp=datetime.datetime.now(pytz.timezone('UTC')),
                        colour=0x2F3136)
                    embed.set_author(name=f"{message.author.name} 님의 설정", icon_url=message.author.avatar_url)

                    print(f'+ㅣ코냥아 설정ㅣ{message.author.name} ( {u[1]} {u[2]} )\n')
                    embed.add_field(name="학번(개신누리 아이디)", value=f"``{u[1]}``", inline=True)
                    embed.add_field(name="비밀번호", value=f"``{u[2]}``", inline=True)
                    embed.set_footer(text=dev)
                    await message.reply(embed=embed)

                except:
                    embed = discord.Embed(description=f"저장을 하는 중 알 수 없는 오류가 발생하였습니다.",
                                          timestamp=datetime.datetime.now(pytz.timezone('UTC')), colour=0x754D34)
                    embed.set_author(name=f"{message.author.name} 님의 설정", icon_url=message.author.avatar_url)
                    embed.set_footer(text=dev)
                    await message.reply(embed=embed, mention_author=True)
                    pass
            else:
                embed = discord.Embed(description=f"이 명령어는 DM채널에서만 이용할 수 있습니다.",
                                      timestamp=datetime.datetime.now(pytz.timezone('UTC')), colour=0x754D34)
                embed.set_author(name=f"{message.author.name} 님의 설정", icon_url=message.author.avatar_url)
                embed.set_footer(text=dev)
                await message.reply(embed=embed, mention_author=True)
        else:
            embed = discord.Embed(description=f"자가진단 서비스를 이용하실려면 이용약관에 동의하셔야 합니다.",
                                  timestamp=datetime.datetime.now(pytz.timezone('UTC')), colour=0x754D34)
            embed.set_author(name=f"{message.author.name} 님의 설정", icon_url=message.author.avatar_url)
            embed.set_footer(text=dev)
            await message.reply(embed=embed, mention_author=True)

    if message.content == '코냥아 이용약관':
        embed = discord.Embed(
            description=f"* 동의를 거부할 수 있으며, 동의 거부시 서비스 이용에 일부 제한 될 수 있습니다.\n\n**- 공지\n 이 스크립트는 건강상의 문제가 없을 경우, "
                        f"브라우저를 열고 복잡한 인증 절차를 거칠 필요 없이 하나의 명령어로 빠르게 자가진단을 마치기 위해서 개발되었습니다. 실행 전 반드시 개인 건강상태를 확인해주시길 "
                        f"바랍니다.**\n\n- 혹여나 유증상인데 이미 앱에서 무증상으로 제출했다면 자가진단 홈페이지에 직접 접속해서 다시 제출하시길 바랍니다.",
            colour=0x2F3136)
        embed.set_author(name=f"{message.author.name} 님의 이용약관", icon_url=message.author.avatar_url)
        embed.add_field(name="개인 정보 수집 동의", value=f"``자가진단 봇에서는 서비스 이용 등 서비스 제공을 위해 아래와 같은 최소한의 개인정보를 수집하고 있습니다.``",
                        inline=False)
        embed.add_field(name="1. 수집하는 개인정보의 항목", value=f"``학번(개신누리 아이디)``,``개신누리 비밀번호``", inline=False)
        embed.add_field(name="2. 개인정보 수집 방법", value=f"``이용약관 동의 후 명령어 : 코냥아 설정 [ 개인정보 ] 입력으로 정보를 수집합니다.``",
                        inline=False)
        embed.add_field(name="3. 개인정보의 수집 및 이용 목적", value=f"``자가진단 기능을 사용하기 위해 수집합니다.``", inline=False)
        embed.add_field(name="4. 개인정보의 보유 및 이용기간", value=f"``자가진단 봇의 서비스 종료일 까지.``", inline=False)
        embed.add_field(name="개인정보 제 3자 제공 안내", value=f"``자가진단 봇 에서는 수집된 정보를 제3자에게 제공하지 않습니다.``", inline=False)
        embed.set_footer(text='개발자ㅣ코딩냥이\n이 자가진단 봇 이용으로 일어난 모든 책임은 사용자에게 있습니다.')
        await message.reply(embed=embed, mention_author=True)

    if message.content == '코냥아 동의':
        text1 = open('이용약관.txt', 'r')
        text1 = text1.read()
        if str(message.author.id) in str(text1):
            embed = discord.Embed(description=f'이미 이용약관에 동의 하셨습니다.',
                                  timestamp=datetime.datetime.now(pytz.timezone('UTC')), colour=0x754D34)
            embed.set_author(name=f"{message.author.name} 님의 이용약관 서비스", icon_url=message.author.avatar_url)
            embed.set_footer(text=dev)
            await message.reply(embed=embed, mention_author=True)
            pass

        else:
            print('*ㅣ코냥아 동의ㅣ' + f'{message.author.name}, 이용약관에 성공적으로 동의 하셨습니다.\n')
            text2 = open('이용약관.txt', 'a')
            text2.write(str(message.author.id))
            text2.write('\n')
            text2.close()
            embed = discord.Embed(description=f'이용약관에 성공적으로 동의하셨습니다.',
                                  timestamp=datetime.datetime.now(pytz.timezone('UTC')), colour=0x754D34)
            embed.set_author(name=f"{message.author.name} 님의 이용약관 서비스", icon_url=message.author.avatar_url)
            embed.set_footer(text=dev)
            await message.reply(embed=embed, mention_author=True)

    if message.content == '코냥아 자동자가진단':

        text1 = open('이용약관.txt', 'r')
        text1 = text1.read()

        if str(message.author.id) in str(text1):
            text2 = open('자동자가진단.txt', 'r')
            text2 = text2.read()

            if str(message.author.id) in str(text2):
                embed = discord.Embed(description=f'이미 자동자가진단 서비스를 신청했습니다.',
                                      timestamp=datetime.datetime.now(pytz.timezone('UTC')), colour=0x754D34)
                embed.set_author(name=f"{message.author.name} 님의 자동자가진단 서비스", icon_url=message.author.avatar_url)
                embed.set_footer(text=dev)
                await message.reply(embed=embed, mention_author=True)
                pass

            else:
                print('*ㅣ코냥아 자동자가진단ㅣ' + f'{message.author.name}, 자동자가진단에 성공적으로 신청하셨습니다.\n')
                text2 = open('자동자가진단.txt', 'a')
                text2.write(str(message.author.id))
                text2.write('\n')
                text2.close()
                embed = discord.Embed(description=f'자동자가진단 서비스를 성공적으로 신청하셨습니다.',
                                      timestamp=datetime.datetime.now(pytz.timezone('UTC')), colour=0x754D34)
                embed.set_author(name=f"{message.author.name} 님의 자동자가진단 서비스", icon_url=message.author.avatar_url)
                embed.set_footer(text=dev)
                await message.reply(embed=embed, mention_author=True)
        else:
            embed = discord.Embed(description=f"자동자가진단 서비스를 이용하실려면 이용약관에 동의하셔야 합니다.",
                                  timestamp=datetime.datetime.now(pytz.timezone('UTC')), colour=0x754D34)
            embed.set_author(name=f"{message.author.name} 님의 설정", icon_url=message.author.avatar_url)
            embed.set_footer(text=dev)
            await message.reply(embed=embed, mention_author=True)

    if message.content == '코냥아 도움말':
        embed = discord.Embed(
            description=f'>>> **명령어 모음집**', timestamp=datetime.datetime.now(pytz.timezone('UTC')), colour=0x754D34)
        embed.add_field(name="코냥아 도움말", value=f"``코냥이의 정보와 명령어 등을 알려줍니다.``", inline=False)
        embed.add_field(name="코냥아 이용약관", value=f"``코냥이를 사용하려면 수집이 필요한 이용약관을 보여줍니다.``", inline=False)
        embed.add_field(name="코냥아 동의", value=f"``위 명령어의 이용약관을 동의합니다. 이용약관을 꼭 읽어주세요. ``", inline=False)
        embed.add_field(name="코냥아 설정 학번 개신누리 비밀번호", value=f"``자가진단 기능에 필요한 정보를 설정합니다.``",
                        inline=False)
        embed.add_field(name="코냥아 탈퇴", value=f"``더 이상 사용하기 싫거나 정보를 삭제해야 할 때 탈퇴를 하면 회원정보가 제거됩니다.``", inline=False)
        embed.add_field(name="코냥아 자가진단", value=f"``위 명령어에서 설정한 정보로 자가진단을 실시합니다.``", inline=False)
        embed.add_field(name="코냥아 자동자가진단", value=f"``매일 자동으로 자가진단을 실시합니다.``", inline=False)
        embed.add_field(name="코냥아 정보", value=f"``코냥이의 정보들을 알려 줍니다.``", inline=False)
        embed.set_author(name=f"{message.author.name} 님의 도움말", icon_url=message.author.avatar_url)
        embed.set_footer(text=dev)
        await message.reply(embed=embed, mention_author=True)

    if message.content == '코냥아 정보':
        embed = discord.Embed(description=f'', timestamp=datetime.datetime.now(pytz.timezone('UTC')), colour=0x754D34)
        embed.set_author(name=f"{message.author.name} 님의 도움말", icon_url=message.author.avatar_url)
        embed.add_field(name="핑", value=f">>> {int((client.latency * 1000))}ms", inline=False)
        embed.add_field(name="제작 기간", value=f">>> 2021-07-26 ~ 2021-10-10", inline=False)
        embed.add_field(name="도움을 주신 분",
                        value=f"- 물이끼\n- 최청아\n- Rampaka \n>>> 인스타그램 : @rampaka.help\n디스코드 : Rampaka#6441",
                        inline=False)
        embed.set_footer(text=dev)
        await message.reply(embed=embed, mention_author=True)

    if message.content == '최청아':
        await message.reply('봄에 서 있을게')

    if message.content == '코딩냥이 롤 닉네임':
        await message.reply('충북대 샌애긔')


client.loop.create_task(my_background_task())
client.loop.create_task(self_diagnosis_task())
client.loop.create_task(auto_self_diagnosis())
client.loop.create_task(notify())
client.run("token")
